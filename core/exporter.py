from typing import List, Dict, Any
from datetime import datetime
import os
from core.opencv_detector import DetectionResult


class ResultExporter:
    """
    Export detection results to Excel format.
    Tracks detection history and provides basic statistics.
    """
    
    def __init__(self):
        self.results_history = []
    
    def add_results(self, results: List[DetectionResult], source: str, timestamp: str = None):
        """
        Add detection results to export queue.
        Only store unique content (data) for Excel export.
        """
        for result in results:
            if not any(r['data'] == result.data for r in self.results_history):
                self.results_history.append({'data': result.data})
    
    # Excel export is the only export method we need
    
    def export_to_excel(self, filename: str, results: List[Dict] = None) -> bool:
        """
        Export results to Excel format with multiple sheets using openpyxl directly.
        
        Args:
            filename: Output Excel filename
            results: Optional specific results to export, uses history if None
            
        Returns:
            True if export successful, False otherwise
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            data_to_export = results if results is not None else self.results_history
            
            if not data_to_export:
                return False
            
            # Ensure directory exists
            os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else '.', exist_ok=True)
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create Detection Results sheet
            ws_results = wb.create_sheet("Detection_Results")
            
            # Headers - Simplified to only include S.No. and Content
            headers = ['S.No.', 'Content']
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col, header in enumerate(headers, 1):
                cell = ws_results.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add data - Simplified to only include serial number and content
            for row, item in enumerate(data_to_export, 2):
                ws_results.cell(row=row, column=1, value=row-1)  # S.No. starts from 1
                ws_results.cell(row=row, column=2, value=item.get('data', ''))
            
            # Auto-adjust column widths for better readability
            ws_results.column_dimensions['A'].width = 10  # S.No. column
            ws_results.column_dimensions['B'].width = 80  # Content column
            
            # Create a simple Summary sheet
            ws_summary = wb.create_sheet("Summary")
            
            # Headers for summary
            ws_summary.cell(row=1, column=1, value="Metric").font = header_font
            ws_summary.cell(row=1, column=1).fill = header_fill
            ws_summary.cell(row=1, column=2, value="Value").font = header_font
            ws_summary.cell(row=1, column=2).fill = header_fill
            
            # Add basic summary data
            ws_summary.cell(row=2, column=1, value="Total Detections")
            ws_summary.cell(row=2, column=2, value=len(data_to_export))
            
            ws_summary.cell(row=3, column=1, value="Date/Time")
            ws_summary.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
            # Auto-adjust summary columns
            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 20
            
            # Save workbook
            wb.save(filename)
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False
    
    # We only need the Excel export and statistics functionality
    
    def _generate_summary_stats(self, data: List[Dict]) -> Dict[str, Any]:
        """
        Generate summary statistics from detection data.
        
        Args:
            data: List of detection result dictionaries
            
        Returns:
            Dictionary of summary statistics
        """
        if not data:
            return {}
        
        # Calculate statistics manually
        confidences = [item.get('confidence', 0) for item in data]
        detection_times = [item.get('detection_time', 0) for item in data]
        sources = [item.get('source', '') for item in data]
        types = [item.get('type', '') for item in data]
        timestamps = [item.get('timestamp', '') for item in data]
        
        # Count unique values
        unique_sources = len(set(sources))
        unique_types = len(set(types))
        
        # Calculate averages
        avg_confidence = sum(confidences) / len(confidences) if confidences else 0
        avg_time = sum(detection_times) / len(detection_times) if detection_times else 0
        
        # Find min/max detection times
        min_time = min(detection_times) if detection_times else 0
        max_time = max(detection_times) if detection_times else 0
        
        # Find most common type
        type_counts = {}
        for t in types:
            type_counts[t] = type_counts.get(t, 0) + 1
        most_common_type = max(type_counts.items(), key=lambda x: x[1])[0] if type_counts else 'N/A'
        
        # Date range
        valid_timestamps = [t for t in timestamps if t]
        date_range = f"{min(valid_timestamps)} to {max(valid_timestamps)}" if valid_timestamps else "N/A"
        
        stats = {
            'Total Detections': len(data),
            'Unique Sources': unique_sources,
            'Code Types Detected': unique_types,
            'Average Confidence': f"{avg_confidence:.3f}",
            'Average Detection Time': f"{avg_time:.3f}s",
            'Min Detection Time': f"{min_time:.3f}s",
            'Max Detection Time': f"{max_time:.3f}s",
            'Most Common Type': most_common_type,
            'Date Range': date_range
        }
        
        # Add type counts
        for code_type, count in type_counts.items():
            stats[f'{code_type} Count'] = count
        
        return stats
    
    def clear_history(self):
        """Clear all stored results."""
        self.results_history.clear()
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get current statistics without exporting."""
        return self._generate_summary_stats(self.results_history)
